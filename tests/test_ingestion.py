from pathlib import Path

from openpyxl import Workbook

from aifaq import ingestion


def _make_repo(tmp_path: Path) -> Path:
    (tmp_path / "knowledge" / "public").mkdir(parents=True)
    return tmp_path


def test_excel_multi_sheet_and_hidden_sheet_excluded(tmp_path, settings, repos):
    repo_root = _make_repo(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "申請先"
    ws.append(["項目", "値"])
    ws.append(["対象", "検査PC交換"])
    hidden = wb.create_sheet("hidden")
    hidden.sheet_state = "hidden"
    hidden.append(["secret"])
    path = repo_root / "knowledge" / "public" / "sample.xlsx"
    wb.save(path)

    outcome = ingestion.import_file(repos, settings, repo_root, path)
    assert outcome.status == "imported"

    sheets = {
        row["sheet_name"]
        for row in repos.conn.execute("SELECT DISTINCT sheet_name FROM source_chunks").fetchall()
    }
    assert "申請先" in sheets
    assert "hidden" not in sheets


def test_xlsm_does_not_execute_macro_only_reads_values(tmp_path, settings, repos):
    repo_root = _make_repo(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append(["1", "2"])
    path = repo_root / "knowledge" / "public" / "sample.xlsm"
    wb.save(path)

    outcome = ingestion.import_file(repos, settings, repo_root, path)
    assert outcome.status == "imported"
    rows = repos.conn.execute("SELECT content FROM source_chunks").fetchall()
    assert any("1" in r["content"] for r in rows)


def test_csv_utf8_and_cp932_fallback(tmp_path, settings, repos):
    repo_root = _make_repo(tmp_path)
    utf8_path = repo_root / "knowledge" / "public" / "utf8.csv"
    utf8_path.write_text("name,desc\nWi-Fi,無線LAN\n", encoding="utf-8")
    outcome = ingestion.import_file(repos, settings, repo_root, utf8_path)
    assert outcome.status == "imported"
    assert not any("CP932" in w.message for w in outcome.warnings)

    cp932_path = repo_root / "knowledge" / "public" / "cp932.csv"
    cp932_path.write_bytes("name,desc\nプリンタ,故障\n".encode("cp932"))
    outcome2 = ingestion.import_file(repos, settings, repo_root, cp932_path)
    assert outcome2.status == "imported"
    assert any("CP932" in w.message for w in outcome2.warnings)


def test_markdown_heading_split(tmp_path, settings, repos):
    repo_root = _make_repo(tmp_path)
    path = repo_root / "knowledge" / "public" / "doc.md"
    path.write_text("# 見出し1\n本文1\n## 見出し2\n本文2\n", encoding="utf-8")
    ingestion.import_file(repos, settings, repo_root, path)
    headings = {
        r["heading"] for r in repos.conn.execute("SELECT heading FROM source_chunks").fetchall()
    }
    assert "見出し1" in headings
    assert "見出し2" in headings


def test_txt_size_limited_chunking(tmp_path, settings, repos):
    repo_root = _make_repo(tmp_path)
    path = repo_root / "knowledge" / "public" / "big.txt"
    path.write_text("あ" * 10000, encoding="utf-8")
    outcome = ingestion.import_file(repos, settings, repo_root, path)
    assert outcome.status == "imported"
    rows = repos.conn.execute("SELECT content FROM source_chunks").fetchall()
    assert len(rows) >= 2
    assert all(len(r["content"]) <= settings.max_text_chunk_chars + 1 for r in rows)


def test_sha256_skip_unchanged_and_reimport_changed(tmp_path, settings, repos):
    repo_root = _make_repo(tmp_path)
    path = repo_root / "knowledge" / "public" / "a.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")

    o1 = ingestion.import_file(repos, settings, repo_root, path)
    assert o1.status == "imported"
    o2 = ingestion.import_file(repos, settings, repo_root, path)
    assert o2.status == "skipped_unchanged"

    path.write_text("a,b\n3,4\n", encoding="utf-8")
    o3 = ingestion.import_file(repos, settings, repo_root, path)
    assert o3.status == "imported"


def test_deleted_file_marked_missing(tmp_path, settings, repos):
    repo_root = _make_repo(tmp_path)
    path = repo_root / "knowledge" / "public" / "a.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")
    ingestion.import_all(repos, settings, repo_root)
    path.unlink()
    _, missing = ingestion.import_all(repos, settings, repo_root)
    assert "knowledge/public/a.csv" in missing
    row = repos.source_files.get_by_path("knowledge/public/a.csv")
    assert row["status"] == "missing"
