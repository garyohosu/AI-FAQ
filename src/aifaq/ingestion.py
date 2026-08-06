"""Excel/CSV/TSV/Markdown/テキストの知識資料取り込み。

- Excelは `openpyxl` を `read_only=True`, `data_only=True` で開き、セル値だけを
  読む。openpyxlはVBA/マクロを実行するエンジンではないため、この読み方で
  マクロが実行されることはない。非表示シートは既定で除外する。
- CSV/TSV/テキストはUTF-8を優先し、失敗時のみCP932へフォールドする。
- 相対パスの安全性は `models.SourceFileRecord`/`SourceIndexEntry` の
  バリデータで検証される (`..`・絶対パス・UNCパスを拒否)。
"""

from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from aifaq.config import Settings
from aifaq.models import (
    ImportWarning,
    LocationType,
    SourceChunk,
    SourceFileRecord,
    SourcePriority,
    SourceScope,
    SourceStatus,
)
from aifaq.repositories import Repositories

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv", ".tsv", ".md", ".txt"}

_SCOPE_BY_FOLDER = {
    "public": SourceScope.PUBLIC,
    "internal": SourceScope.INTERNAL,
    "policies": SourceScope.INTERNAL,
    "procedures": SourceScope.INTERNAL,
    "troubleshooting": SourceScope.INTERNAL,
    "inventory": SourceScope.INTERNAL,
    "inbox": SourceScope.INTERNAL,
}


@dataclass
class FileImportOutcome:
    relative_path: str
    status: str  # imported | skipped_unchanged | failed
    warnings: list[ImportWarning]


def sha256_of_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(65536), b""):
            h.update(block)
    return h.hexdigest()


def content_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _default_scope(relative_path: str) -> SourceScope:
    parts = Path(relative_path).parts
    if len(parts) >= 2 and parts[0] == "knowledge":
        return _SCOPE_BY_FOLDER.get(parts[1], SourceScope.INTERNAL)
    return SourceScope.INTERNAL


def scan_supported_files(settings: Settings, repo_root: Path) -> list[Path]:
    """`knowledge/` 配下の対応形式ファイルを列挙する(内容は読まない)。"""
    knowledge_dir = repo_root / settings.knowledge_dir
    if not knowledge_dir.exists():
        return []
    files = []
    for path in sorted(knowledge_dir.rglob("*")):
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(path)
    return files


def _read_text_with_fallback(path: Path) -> tuple[str, str]:
    """UTF-8を優先し、失敗時のみCP932へフォールバックする。使用した符号化名を返す。"""
    raw = path.read_bytes()
    try:
        return raw.decode("utf-8"), "utf-8"
    except UnicodeDecodeError:
        return raw.decode("cp932", errors="replace"), "cp932"


def _truncate(text: str, max_chars: int) -> str:
    return text if len(text) <= max_chars else text[:max_chars] + "…"


def _row_to_content(headers: list[str], values: list, max_cell_chars: int) -> str:
    parts = []
    for idx, value in enumerate(values):
        if value is None or str(value).strip() == "":
            continue
        header = headers[idx] if idx < len(headers) and headers[idx] else f"col{idx + 1}"
        parts.append(f"{header}: {_truncate(str(value), max_cell_chars)}")
    return " | ".join(parts)


def read_excel_chunks(
    path: Path, settings: Settings
) -> tuple[list[SourceChunk], list[str]]:
    """openpyxlでExcelを読み、シート・行単位のチャンクを返す。

    read_only=True/data_only=True で開くため、数式・外部リンク・マクロは
    実行されず、保存済みのセル値だけを取得する。
    """
    warnings: list[str] = []
    chunks: list[SourceChunk] = []
    wb = load_workbook(
        filename=str(path), read_only=True, data_only=True, keep_links=False
    )
    try:
        for sheet in wb.worksheets:
            if sheet.sheet_state != "visible":
                continue
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            headers = [str(h) if h is not None else "" for h in header_row][
                : settings.max_import_cols
            ]
            row_idx = 1
            processed = 0
            for row in rows_iter:
                row_idx += 1
                if processed >= settings.max_import_rows:
                    warnings.append(
                        f"シート「{sheet.title}」: 行数上限({settings.max_import_rows})に"
                        "達したため以降の行を取り込みませんでした"
                    )
                    break
                values = list(row)[: settings.max_import_cols]
                content = _row_to_content(headers, values, settings.max_cell_chars)
                if not content:
                    continue
                chunks.append(
                    SourceChunk(
                        source_file_id=0,
                        location_type=LocationType.SHEET_ROWS,
                        sheet_name=sheet.title,
                        row_start=row_idx,
                        row_end=row_idx,
                        heading=None,
                        content=content,
                        content_hash=content_hash(content),
                    )
                )
                processed += 1
    finally:
        wb.close()
    return chunks, warnings


def read_delimited_chunks(
    path: Path, settings: Settings, delimiter: str
) -> tuple[list[SourceChunk], list[str]]:
    warnings: list[str] = []
    text, encoding = _read_text_with_fallback(path)
    if encoding != "utf-8":
        warnings.append(f"文字コードをCP932と推定して読み込みました: {path.name}")
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return [], warnings
    headers = [h.strip() for h in rows[0]][: settings.max_import_cols]
    chunks: list[SourceChunk] = []
    for row_idx, values in enumerate(rows[1:], start=2):
        if row_idx - 1 > settings.max_import_rows:
            warnings.append(
                f"行数上限({settings.max_import_rows})に達したため以降の行を"
                "取り込みませんでした"
            )
            break
        content = _row_to_content(headers, values[: settings.max_import_cols], settings.max_cell_chars)
        if not content:
            continue
        chunks.append(
            SourceChunk(
                source_file_id=0,
                location_type=LocationType.CSV_ROWS,
                sheet_name=None,
                row_start=row_idx,
                row_end=row_idx,
                heading=None,
                content=content,
                content_hash=content_hash(content),
            )
        )
    return chunks, warnings


def read_markdown_chunks(
    path: Path, settings: Settings
) -> tuple[list[SourceChunk], list[str]]:
    warnings: list[str] = []
    text, encoding = _read_text_with_fallback(path)
    if encoding != "utf-8":
        warnings.append(f"文字コードをCP932と推定して読み込みました: {path.name}")

    lines = text.splitlines()
    sections: list[tuple[str | None, int, list[str]]] = []
    current_heading: str | None = None
    current_start = 1
    current_lines: list[str] = []

    def flush(end_line: int) -> None:
        if current_lines:
            sections.append((current_heading, current_start, current_lines[:]))

    for i, line in enumerate(lines, start=1):
        if line.startswith("#"):
            flush(i - 1)
            current_lines.clear()
            current_heading = line.lstrip("#").strip()
            current_start = i
        else:
            current_lines.append(line)
    flush(len(lines))

    chunks: list[SourceChunk] = []
    for heading, start, body_lines in sections:
        content = "\n".join(body_lines).strip()
        if not content:
            continue
        content = _truncate(content, settings.max_text_chunk_chars)
        chunks.append(
            SourceChunk(
                source_file_id=0,
                location_type=LocationType.MARKDOWN_HEADING,
                sheet_name=None,
                row_start=start,
                row_end=start + len(body_lines),
                heading=heading,
                content=content,
                content_hash=content_hash(content),
            )
        )
    return chunks, warnings


def read_text_chunks(
    path: Path, settings: Settings
) -> tuple[list[SourceChunk], list[str]]:
    """`max_text_chunk_chars` 文字ごとの固定長ブロックに分割する。

    改行の有無に関わらず等分割するため、改行を含まない巨大な1行の
    テキストでも内容を欠落させない。
    """
    warnings: list[str] = []
    text, encoding = _read_text_with_fallback(path)
    if encoding != "utf-8":
        warnings.append(f"文字コードをCP932と推定して読み込みました: {path.name}")

    max_chars = settings.max_text_chunk_chars
    chunks: list[SourceChunk] = []
    pos = 0
    n = len(text)
    while pos < n:
        end = min(pos + max_chars, n)
        block = text[pos:end].strip()
        if block:
            row_start = text.count("\n", 0, pos) + 1
            row_end = text.count("\n", 0, end) + 1
            chunks.append(
                SourceChunk(
                    source_file_id=0,
                    location_type=LocationType.TEXT_BLOCK,
                    content=block,
                    content_hash=content_hash(block),
                    row_start=row_start,
                    row_end=row_end,
                )
            )
        pos = end
    return chunks, warnings


def import_file(
    conn_repos: Repositories,
    settings: Settings,
    repo_root: Path,
    path: Path,
    *,
    scope_override: SourceScope | None = None,
    category: str = "other",
    owner: str | None = None,
    priority: SourcePriority = SourcePriority.NORMAL,
    status: SourceStatus = SourceStatus.ACTIVE,
) -> FileImportOutcome:
    relative_path = path.relative_to(repo_root).as_posix()
    warnings: list[ImportWarning] = []

    size_bytes = path.stat().st_size
    if size_bytes > settings.max_file_size_bytes:
        warnings.append(
            ImportWarning(
                file=relative_path,
                message=f"ファイルサイズが上限({settings.max_file_size_bytes} bytes)を超過",
                severity="failed",
            )
        )
        return FileImportOutcome(relative_path, "failed", warnings)

    digest = sha256_of_file(path)
    existing = conn_repos.source_files.get_by_path(relative_path)
    if existing is not None and existing["sha256"] == digest and existing["status"] == "active":
        return FileImportOutcome(relative_path, "skipped_unchanged", warnings)

    ext = path.suffix.lower()
    try:
        if ext == ".xlsx" or ext == ".xlsm":
            chunks, msgs = read_excel_chunks(path, settings)
        elif ext == ".csv":
            chunks, msgs = read_delimited_chunks(path, settings, delimiter=",")
        elif ext == ".tsv":
            chunks, msgs = read_delimited_chunks(path, settings, delimiter="\t")
        elif ext == ".md":
            chunks, msgs = read_markdown_chunks(path, settings)
        elif ext == ".txt":
            chunks, msgs = read_text_chunks(path, settings)
        else:
            warnings.append(
                ImportWarning(file=relative_path, message="未対応の拡張子", severity="failed")
            )
            return FileImportOutcome(relative_path, "failed", warnings)
    except Exception as exc:  # noqa: BLE001 - 1ファイルの失敗で全体を止めない
        warnings.append(
            ImportWarning(file=relative_path, message=f"読み込み失敗: {exc}", severity="failed")
        )
        record = SourceFileRecord(
            relative_path=relative_path,
            file_type=ext.lstrip("."),
            scope=scope_override or _default_scope(relative_path),
            category=category,
            owner=owner,
            priority=priority,
            status=SourceStatus.ACTIVE,
            size_bytes=size_bytes,
            modified_at=datetime.fromtimestamp(path.stat().st_mtime, tz=UTC),
            sha256=digest,
            error_summary=str(exc)[:500],
        )
        conn_repos.source_files.upsert(record)
        return FileImportOutcome(relative_path, "failed", warnings)

    warnings.extend(ImportWarning(file=relative_path, message=m) for m in msgs)

    record = SourceFileRecord(
        relative_path=relative_path,
        file_type=ext.lstrip("."),
        scope=scope_override or _default_scope(relative_path),
        category=category,
        owner=owner,
        priority=priority,
        status=status,
        size_bytes=size_bytes,
        modified_at=datetime.fromtimestamp(path.stat().st_mtime, tz=UTC),
        sha256=digest,
        last_imported_at=datetime.now(UTC),
        error_summary=None,
    )
    source_file_id = conn_repos.source_files.upsert(record)
    for chunk in chunks:
        chunk.source_file_id = source_file_id
    conn_repos.source_chunks.replace_for_file(source_file_id, chunks)
    return FileImportOutcome(relative_path, "imported", warnings)


def import_all(
    conn_repos: Repositories,
    settings: Settings,
    repo_root: Path,
    *,
    only_path: Path | None = None,
) -> tuple[list[FileImportOutcome], list[str]]:
    """`knowledge/` を走査し、差分だけ取り込む。1ファイルの失敗で全体を止めない。"""
    files = [only_path] if only_path else scan_supported_files(settings, repo_root)
    outcomes: list[FileImportOutcome] = []
    for path in files:
        outcomes.append(import_file(conn_repos, settings, repo_root, path))

    missing_notes: list[str] = []
    if only_path is None:
        on_disk = {p.relative_to(repo_root).as_posix() for p in files}
        for row in conn_repos.source_files.list():
            if row["relative_path"] not in on_disk and row["status"] != "missing":
                conn_repos.source_files.mark_missing(row["relative_path"])
                missing_notes.append(row["relative_path"])
    return outcomes, missing_notes
